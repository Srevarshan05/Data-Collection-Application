"""
Utility functions for image processing, validation, and reporting
"""

import os
import re
from PIL import Image
from datetime import datetime, timedelta
import pandas as pd
from typing import Optional, Tuple


# Registration number prefixes based on year
YEAR_PREFIXES = {
    1: "RA2511026050",
    2: "RA2411026050",
    3: "RA2311026050"
}

# Valid sections per year
YEAR_SECTIONS = {
    1: ["A", "B", "C", "D", "E"],
    2: ["A", "B", "C", "D", "E"],
    3: ["A", "B", "C", "D"]
}

# Image settings
IMAGE_SIZE = (300, 300)
IMAGE_QUALITY = 70
MAX_FILE_SIZE = 500 * 1024  # 500KB in bytes
ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png'}


def get_registration_prefix(year: int) -> Optional[str]:
    """
    Get registration number prefix for a given year
    """
    return YEAR_PREFIXES.get(year)


def validate_year_section(year: int, section: str) -> bool:
    """
    Validate if section is valid for the given year
    """
    valid_sections = YEAR_SECTIONS.get(year, [])
    return section.upper() in valid_sections


def validate_last_digits(last_digits: str) -> bool:
    """
    Validate that last digits are exactly 3 numeric characters
    """
    return bool(re.match(r'^\d{3}$', last_digits))


def generate_register_number(year: int, last_digits: str) -> Optional[str]:
    """
    Generate complete registration number
    """
    prefix = get_registration_prefix(year)
    if not prefix:
        return None
    return f"{prefix}{last_digits}"


def validate_file_extension(filename: str) -> bool:
    """
    Validate file extension
    """
    ext = os.path.splitext(filename)[1].lower()
    return ext in ALLOWED_EXTENSIONS


def validate_file_size(file_size: int) -> bool:
    """
    Validate file size (max 500KB)
    """
    return file_size <= MAX_FILE_SIZE


def create_upload_directory(year: int, section: str) -> str:
    """
    Create and return upload directory path
    """
    upload_dir = os.path.join("uploads", str(year), section.upper())
    os.makedirs(upload_dir, exist_ok=True)
    return upload_dir


def process_and_save_image(image_file, year: int, section: str, register_number: str) -> str:
    """
    Process image: resize, compress, and save
    Returns the saved file path
    """
    # Create directory
    upload_dir = create_upload_directory(year, section)
    
    # Generate filename
    filename = f"{register_number}.jpg"
    filepath = os.path.join(upload_dir, filename)
    
    # Open and process image
    img = Image.open(image_file)
    
    # Convert to RGB if necessary (for PNG with transparency)
    if img.mode in ('RGBA', 'LA', 'P'):
        background = Image.new('RGB', img.size, (255, 255, 255))
        if img.mode == 'P':
            img = img.convert('RGBA')
        background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
        img = background
    
    # Resize image to 300x300
    img = img.resize(IMAGE_SIZE, Image.Resampling.LANCZOS)
    
    # Save with compression
    img.save(filepath, 'JPEG', quality=IMAGE_QUALITY, optimize=True)
    
    return filepath


def get_file_size(file) -> int:
    """
    Get file size in bytes
    """
    file.seek(0, 2)  # Seek to end
    size = file.tell()
    file.seek(0)  # Reset to beginning
    return size


def generate_csv_report(students_data: list, filename: str = None) -> str:
    """
    Generate CSV report from student data
    """
    # Create reports directory if it doesn't exist
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)
    
    # Generate filename if not provided
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"student_report_{timestamp}.csv"
    
    filepath = os.path.join(reports_dir, filename)
    
    # Create DataFrame
    df = pd.DataFrame(students_data)
    
    # Reorder columns
    column_order = ['name', 'year', 'section', 'register_number', 'created_at']
    df = df[column_order]
    
    # Rename columns for better readability
    df.columns = ['Name', 'Year', 'Section', 'Register Number', 'Registration Date']
    
    # Save to CSV
    df.to_csv(filepath, index=False)
    
    return filepath


def get_weekly_registrations(students_data: list) -> list:
    """
    Filter students registered in the last 7 days
    """
    week_ago = datetime.now() - timedelta(days=7)
    
    weekly_students = []
    for student in students_data:
        created_at = student.get('created_at')
        if created_at:
            # Parse datetime string
            if isinstance(created_at, str):
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            
            if created_at.replace(tzinfo=None) >= week_ago:
                weekly_students.append(student)
    
    return weekly_students


def get_year_wise_count(students_data: list) -> dict:
    """
    Get count of students per year
    """
    year_count = {1: 0, 2: 0, 3: 0}
    
    for student in students_data:
        year = student.get('year')
        if year in year_count:
            year_count[year] += 1
    
    return year_count


def get_section_wise_count(students_data: list) -> dict:
    """
    Get count of students per section
    """
    section_count = {}
    
    for student in students_data:
        section = student.get('section')
        if section:
            section_count[section] = section_count.get(section, 0) + 1
    
    return section_count


def generate_excel_report_with_photos(students_data: list, filename: str = None) -> str:
    """
    Generate Excel report with embedded student photos
    """
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel generation. Install it with: pip install openpyxl")
    
    # Create reports directory if it doesn't exist
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)
    
    # Generate filename if not provided
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"student_report_{timestamp}.xlsx"
    
    # Ensure .xlsx extension
    if not filename.endswith('.xlsx'):
        filename = filename.replace('.csv', '.xlsx')
    
    filepath = os.path.join(reports_dir, filename)
    
    # Create workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Records"
    
    # Define headers
    headers = ['Photo', 'Name', 'Year', 'Section', 'Register Number', 'Registration Date']
    
    # Style for headers
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Photo column
    ws.column_dimensions['B'].width = 25  # Name
    ws.column_dimensions['C'].width = 10  # Year
    ws.column_dimensions['D'].width = 10  # Section
    ws.column_dimensions['E'].width = 20  # Register Number
    ws.column_dimensions['F'].width = 20  # Registration Date
    
    # Set row height for header
    ws.row_dimensions[1].height = 25
    
    # Add student data
    for idx, student in enumerate(students_data, start=2):
        # Set row height for photo (80 pixels = approximately 60 points)
        ws.row_dimensions[idx].height = 80
        
        # Add photo if exists
        photo_path = student.get('photo_path', '')
        if photo_path and os.path.exists(photo_path):
            try:
                # Open and resize image directly for Excel
                img = Image.open(photo_path)
                
                # Resize image to fit in cell (100x100 pixels)
                img_resized = img.resize((100, 100), Image.Resampling.LANCZOS)
                
                # Create temp directory if it doesn't exist
                temp_dir = os.path.join(os.getcwd(), "reports", "temp")
                os.makedirs(temp_dir, exist_ok=True)
                
                # Save temporary resized image with simple filename
                temp_filename = f"img_{idx}.jpg"
                temp_img_path = os.path.join(temp_dir, temp_filename)
                
                # Save the resized image
                img_resized.save(temp_img_path, 'JPEG', quality=85)
                
                # Add image to Excel
                xl_img = XLImage(temp_img_path)
                xl_img.width = 100
                xl_img.height = 100
                
                # Position image in cell A (Photo column)
                cell_ref = f'A{idx}'
                ws.add_image(xl_img, cell_ref)
                    
            except Exception as e:
                # If image fails, write error message
                cell = ws.cell(row=idx, column=1)
                cell.value = "Photo Error"
                cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell = ws.cell(row=idx, column=1)
            cell.value = "No Photo"
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add other student data
        # Name
        cell = ws.cell(row=idx, column=2)
        cell.value = student.get('name', '')
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        
        # Year
        cell = ws.cell(row=idx, column=3)
        cell.value = student.get('year', '')
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        # Section
        cell = ws.cell(row=idx, column=4)
        cell.value = student.get('section', '')
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        # Register Number
        cell = ws.cell(row=idx, column=5)
        cell.value = student.get('register_number', '')
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        
        # Registration Date
        cell = ws.cell(row=idx, column=6)
        created_at = student.get('created_at', '')
        if created_at:
            if isinstance(created_at, str):
                try:
                    dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    cell.value = dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    cell.value = created_at
            else:
                cell.value = created_at.strftime('%Y-%m-%d %H:%M:%S')
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
    
    # Save workbook
    wb.save(filepath)
    
    # Clean up all temporary images after saving Excel
    try:
        temp_dir = os.path.join(os.getcwd(), "reports", "temp")
        if os.path.exists(temp_dir):
            for file in os.listdir(temp_dir):
                if file.startswith("img_") and file.endswith(".jpg"):
                    try:
                        os.remove(os.path.join(temp_dir, file))
                    except:
                        pass
    except:
        pass
    
    return filepath


def format_file_size(size_bytes: int) -> str:
    """
    Format file size in human-readable format
    """
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.2f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.2f} MB"
